#!/usr/bin/python

#
# colour-code halpotype according to haplocode information
# this version loads the data into a numpy array, rather than working line-by-line
# in a file
#

import sys
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

haplotype = sys.argv[1]
haplocode = sys.argv[2]
focal = sys.argv[3]
outfile = sys.argv[4]

f = open(focal)
focalsnps = [x.strip() for x in f]
f.close()

m_htype = np.genfromtxt(haplotype,dtype=object,delimiter=',') #haplotype (affymetrix allele code A or B)
m_hcode = np.genfromtxt(haplocode,dtype=object,delimiter=',') #haplocodes (denotes which of the two parental haplotype is present)

m_htype = np.insert(m_htype, 1, 0, axis=0) #insert new row to help sorting
m_hcode = np.insert(m_hcode, 1, 0, axis=0)
   
wb = Workbook()
ws = wb.get_active_sheet()
ws.title = "ColouredHaplotypes"

n_rows = m_htype.shape[0]
n_cols = m_htype.shape[1]

#interpolate haplocode where data is missing due to marker type
for c in xrange(8,n_cols):
    py = 'x' #previous haplocode
    pr = None #previous row
    for r in xrange(2,n_rows):
        y = m_hcode[r,c]
        if y == '-': continue
        if y == py:
            if pr != r-1:
                for rr in xrange(pr+1,r):
                    m_hcode[rr,c] = y
                            
        py = y
        pr = r

#make dictionary of all snps in this LG
snpdict = {m_htype[i,0]:i for i in xrange(2,n_rows)}

#find the focal snp id and row number
focalid = None
for x in focalsnps:
    if x in snpdict:
        focalid = x
        focalrow = snpdict[x]
        print outfile,x
        break
        
if focalid == None: print outfile, "None"

for i in xrange(0,n_cols):
    m_htype[1,i] = i #initial sort order - no change

#find distance from focal snp to nearest recombination point
if focalid != None:
    for c in xrange(8,n_cols,2):
        mindist = 9999999
        hcode1 = m_hcode[focalrow,c]
        hcode2 = m_hcode[focalrow,c+1]
        for r in xrange(2,n_rows):
            if m_hcode[r,c] != '-' and m_hcode[r,c] != hcode1:
                #print m_hcode[r,c],hcode1
                if abs(focalrow-r) < mindist:
                    mindist = abs(focalrow-r)
            if m_hcode[r,c+1] != '-' and m_hcode[r,c+1] != hcode2:
                if abs(focalrow-r) < mindist:
                    mindist = abs(focalrow-r)
        m_htype[1,c] = 100*mindist
        m_htype[1,c+1] = 100*mindist

#sort m_htype and m_hcode
m_htype = np.transpose(m_htype)
m_hcode = np.transpose(m_hcode)
m_hcode = m_hcode[m_htype[:,1].argsort()]
m_htype = m_htype[m_htype[:,1].argsort()]
m_htype = np.transpose(m_htype)
m_hcode = np.transpose(m_hcode)

fillM0 = PatternFill("solid", fgColor="FF6666")
fillM1 = PatternFill("solid", fgColor="6666FF")
fillP0 = PatternFill("solid", fgColor="66FF66")
fillP1 = PatternFill("solid", fgColor="FFFF66")

#copy header unmodified
for i in xrange(0,n_cols):
    cell = ws.cell(row=1, column=i+1)
    cell.value = m_htype[0,i]
    
    cell = ws.cell(row=2, column=i+1)
    cell.value = m_htype[1,i]
    
#colour code haplotype alleles according to haplocode
for r in xrange(2,n_rows):
    for c in xrange(0,n_cols):
        x = m_htype[r,c]
        y = m_hcode[r,c]
        cell = ws.cell(row=r+1, column=c+1)
        cell.value = x
        
        if r > 0 and c > 7:
            if y == '0':
                if c%2 == 1:
                    cell.fill = fillM0
                else:
                    cell.fill = fillP0
            elif y == '1':
                if c%2 == 1:
                    cell.fill = fillM1
                else:
                    cell.fill = fillP1
                                        
wb.save(outfile)
